# AUTOGENERATED! DO NOT EDIT! File to edit: nbs/filter.ipynb (unless otherwise specified).

__all__ = ['bump_version', 'get_modtimes', 'formatar_db']

# Cell
import os
from pathlib import Path
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from .query import *
from .constants import console, APP_ANALISE
from fastcore.test import *
from fastcore.script import call_parse, Param, store_true
from pyarrow import ArrowInvalid
from geopy.distance import geodesic
from rich import print

# Cell
def bump_version(version, part=2):
    version = version.split(".")
    version[part] = str(int(version[part]) + 1)
    for i in range(part + 1, 3):
        version[i] = "0"
    return ".".join(version)

# Internal Cell
def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs,
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not Path(filename).is_file():
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs,
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(
        filename, engine="openpyxl", mode="a", if_sheet_exists="replace"
    )

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

# Cell
def get_modtimes(pasta):
    """
    Retorna a data de modificação dos arquivos de dados
    """
    # Pasta
    pasta = Path(pasta)
    if not pasta.is_dir():
        raise FileNotFoundError(f"Pasta {pasta} não encontrada")
    # Arquivos
    if not (stel := pasta / 'stel.fth').is_file():
        raise FileNotFoundError(f"Arquivo {stel} não encontrado")
    if not (radcom := pasta / 'radcom.fth').is_file():
        raise FileNotFoundError(f"Arquivo {radcom} não encontrado")
    if not (mosaico := pasta / 'mosaico.fth').is_file():
        raise FileNotFoundError(f"Arquivo {mosaico} não encontrado")
    if not (icao := pasta / 'IcaoDB.xlsx').is_file():  # ICAO
        raise FileNotFoundError(f"Arquivo {icao} não encontrado")
    if not (pmec := pasta / 'PmecDB.xlsx').is_file():  # PMEC
        raise FileNotFoundError(f"Arquivo {pmec} não encontrado")
    if not (geo := pasta / 'GeoAiswebDB.xlsx').is_file():  # GEO
        raise FileNotFoundError(f"Arquivo {geo} não encontrado")
    # Modificação
    mod_stel = datetime.fromtimestamp(stel.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S")
    mod_radcom = datetime.fromtimestamp(radcom.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S")
    mod_mosaico = datetime.fromtimestamp(mosaico.stat().st_mtime).strftime("%d/%m/%Y %H:%M:%S")
    mod_icao = pd.read_excel(icao, engine='openpyxl', sheet_name='ExtractDate').columns[0]
    mod_pmec = pd.read_excel(pmec, engine='openpyxl', sheet_name='ExtractDate').columns[0]
    mod_geo = pd.read_excel(geo, engine='openpyxl', sheet_name='ExtractDate').columns[0]
    return {'STEL': mod_stel,
            'SRD': mod_radcom,
            'MOSAICO': mod_mosaico,
            'ICAO': mod_icao,
            'AISW': mod_pmec,
            'AISG': mod_geo}

@call_parse
def formatar_db(
    path: Param("Pasta onde salvar os arquivos", str),
    up_stel: Param("Atualizar a base do Stel", store_true) = False,
    up_radcom: Param("Atualizar a base do Radcom", store_true) = False,
    up_mosaico: Param("Atualizar a base do Mosaico", store_true) = False,
    up_icao: Param("Atualizar a base do ICAO", store_true) = False,
    up_pmec: Param("Atualizar a base do PMEC", store_true) = False,
    up_geo: Param("Atualizar a base do Geo", store_true) = False,
) -> None:
    dest = Path(path)
    dest.mkdir(parents=True, exist_ok=True)
    console.print(":scroll:[green]Lendo as bases de dados...")
    rd = read_base(path, up_stel, up_radcom, up_mosaico, up_icao)
    rd["Status"] = rd.Status.astype("string")
    rd["Classe"] = rd.Classe.astype("string")
    rd.loc[rd["Classe"].notna(), "Status"] = (
        rd.loc[rd["Classe"].notna(), "Status"]
        + ", "
        + rd.loc[rd["Classe"].notna(), "Classe"]
    )
    rd.loc[rd.Classe == '-1', 'Classe'] = pd.NA
    rd['Classe'] = rd['Classe'].fillna('')
    rd["Descrição"] = (
        "["
        + rd.Fonte.astype("string")
        + "] "
        + rd.Status.astype("string").fillna("-")
        + ", "
        + rd.Entidade.astype("string").fillna("-").str.title()
        + " ("
        + rd.Fistel.astype("string").fillna("-")
        + ", "
        + rd["Número_da_Estação"].astype("string").fillna("-")
        + "), "
        + rd.Município.astype("string").fillna("-")
        + "/"
        + rd.UF.astype("string").fillna("-")
    )


    export_columns = [
        "Frequência",
        "Latitude",
        "Longitude",
        "Descrição",
        "Num_Serviço",
        "Número_da_Estação",
        "Classe_Emissão",
        "BW(kHz)",
    ]
    rd = rd.loc[:, export_columns]
    rd.columns = APP_ANALISE
    common, new = read_aero(path, up_icao, up_pmec, up_geo)
    rd = merge_aero(rd, common, new)
    rd = df_optimize(rd, exclude=["Frequency"])
    rd['Frequency'] = rd['Frequency'].astype('float')
    console.print(":card_file_box:[green]Salvando os arquivos...")
    d = json.loads((dest / "VersionFile.json").read_text())
    mod_times = get_modtimes(path)
    mod_times['ReleaseDate'] = datetime.today().strftime("%d/%m/%Y %H:%M:%S")
    with pd.ExcelWriter(f"{dest}/AnatelDB.xlsx", engine="xlsxwriter") as workbook:
        rd.to_excel(workbook, sheet_name="DataBase", index=False)
    # wb = Workbook()
    # wb.new_sheet('DataBase', data=[rd.columns] + list(rd.values))
    # wb.save(f'{dest}/AnatelDB.xlsx')
    d["anateldb"]["Version"] = bump_version(d["anateldb"]["Version"])
    d['anateldb'].update(mod_times)
    json.dump(d, (dest / "VersionFile.json").open("w"))
    Path(dest / ".version").write_text(f"v{d['anateldb']['Version']}")
    console.print("Sucesso :zap:")
